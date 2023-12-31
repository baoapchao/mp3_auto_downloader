import openpyxl
import pytube as pt
import os
import sys

music_download_folder = 'downloaded_music_files'
music_url_list_excel_file = 'sample_music_url_list.xlsx'

def download_youtube_audios_return_title(url):
    yt = pt.YouTube(url, use_oauth=True, allow_oauth_cache=True)
    t = yt.streams.filter(only_audio=True)
    t[0].download(output_path = music_download_folder)
    return yt.title

def get_excel_value(ws, row, column):
    return ws.cell(row = row, column = column).value



def get_url_list_and_download_audios():
    wb = openpyxl.load_workbook(music_url_list_excel_file)
    ws = wb["Sheet1"]
    for x in range(2, 10000):
        if get_excel_value(ws, x, 1) == None: #No url
            break
        if get_excel_value(ws, x, 7) == 1: #downloaded = 1 then skip to next row
            continue
        title = download_youtube_audios_return_title(get_excel_value(ws, x, 1))
        print('Downloaded', x, title)
        downloaded_cell = ws.cell(row = x, column = 7)
        downloaded_cell.value =  1 #input value: downloaded = 1
        titlecell = ws.cell(row = x, column = 2)
        titlecell.value =  title
        wb.save(music_url_list_excel_file)
        
get_url_list_and_download_audios()

for filename in os.listdir(folder):
    infilename = os.path.join(folder,filename)
    if not os.path.isfile(infilename): continue
    if '.mp3' in filename: continue
    oldbase = os.path.splitext(filename)
    newname = infilename.replace('.mp4', '.mp3')
    output = os.rename(infilename, newname)